"""File parsing for CSV (stdlib) and XLSX (openpyxl) uploads.

Both parsers return a list of row dictionaries keyed by the header row.
Parsing problems raise :class:`ParsingError`.
"""

import csv
import io
from pathlib import PurePosixPath

from openpyxl import load_workbook

from app.core.config import get_settings

settings = get_settings()


class ParsingError(ValueError):
    """Raised when an uploaded file cannot be parsed."""


def _extension(filename: str) -> str:
    """Return the lowercased file extension, guarding against path tricks."""
    # Use only the final path component to avoid directory traversal.
    return PurePosixPath(filename.replace("\\", "/")).suffix.lower()


def validate_upload(filename: str, content: bytes) -> str:
    """Validate filename extension and size; return the normalized extension."""
    extension = _extension(filename)
    if extension not in settings.ALLOWED_UPLOAD_EXTENSIONS:
        raise ParsingError(
            f"Unsupported file type '{extension or 'unknown'}'. "
            f"Allowed: {', '.join(settings.ALLOWED_UPLOAD_EXTENSIONS)}"
        )
    if len(content) > settings.MAX_UPLOAD_SIZE_BYTES:
        raise ParsingError(
            f"File exceeds the maximum allowed size of "
            f"{settings.MAX_UPLOAD_SIZE_BYTES} bytes"
        )
    if not content:
        raise ParsingError("Uploaded file is empty")
    return extension


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ParsingError("File is not valid UTF-8 encoded text") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ParsingError("CSV file has no header row")
    return [{(k or "").strip(): (v or "").strip() for k, v in row.items()} for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various low-level errors
        raise ParsingError("File is not a valid XLSX workbook") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ParsingError("XLSX sheet is empty") from exc

    columns = [str(h).strip() if h is not None else "" for h in header]
    records: list[dict[str, str]] = []
    for raw in rows:
        record = {
            col: ("" if value is None else str(value).strip())
            for col, value in zip(columns, raw)
        }
        records.append(record)
    workbook.close()
    return records


def parse_file(filename: str, content: bytes) -> list[dict[str, str]]:
    """Parse an uploaded file into a list of row dictionaries."""
    extension = validate_upload(filename, content)
    if extension == ".csv":
        return _parse_csv(content)
    return _parse_xlsx(content)
